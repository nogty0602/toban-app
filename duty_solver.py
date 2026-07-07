"""当直表ソルバ本体。solve_roster() を Streamlit から呼ぶ。
見出し行・入局年度行・日直希望行の位置を自動検出する。"""
import io, datetime, statistics
from openpyxl import load_workbook
from ortools.sat.python import cp_model
import jpholiday

FIRST_COL = 4  # A列（メンバー1人目）


def seniority_coeff(year, ymin, ymax, top=2.0):
    if ymax == ymin:
        return 1.0
    return 1.0 + (year - ymin) / (ymax - ymin) * (top - 1.0)


def _norm(v):
    return str(v or "").replace(" ", "").replace("\u3000", "").replace("\n", "").strip()


def _find_layout(ws):
    """見出し行/年度行/希望行 と、日・曜日・当直・氏名開始 の列を自動検出。
    見出し行の中で「日」「曜日」「当直」の各セルを個別に探すので、列がどこでもよい。
    戻り値: (header, year_row, pref_row, c_day, c_wd, c_kind, first_col)"""
    maxc = min(ws.max_column, 15)
    header = c_wd = None
    for r in range(1, min(ws.max_row, 25) + 1):
        for c in range(1, maxc + 1):
            if _norm(ws.cell(r, c).value) == "曜日":
                header, c_wd = r, c
                break
        if header:
            break
    if header is None:
        raise ValueError("見出し行（『曜日』のセル）が見つかりません。"
                         "テンプレに 日/曜日/当直 の見出しがあるかご確認ください。")

    c_day = c_kind = None
    for c in range(1, maxc + 1):
        t = _norm(ws.cell(header, c).value)
        if t == "日" and c_day is None:
            c_day = c
        if ("当直" in t or "当番" in t) and c_kind is None:
            c_kind = c
    if c_day is None:
        c_day = c_wd - 1
    if c_kind is None:
        c_kind = c_wd + 1
    first_col = c_kind + 1

    year_row = pref_row = None
    for r in range(1, min(ws.max_row, 25) + 1):
        for c in range(1, first_col + 1):
            t = _norm(ws.cell(r, c).value)
            if "入局年度" in t:
                year_row = r
            if "希望" in t:
                pref_row = r
    if year_row is None:
        year_row = max(1, header - 1)
    return header, year_row, pref_row, c_day, c_wd, c_kind, first_col


def _read_members(ws, HDR, YR, FC):
    """氏名見出しから (names, years, cols) を列位置で返す（同名・空列があってもOK）。
    3つとも同じ長さのリストで、index が人を一意に表す。"""
    names, years, cols = [], [], []
    for c in range(FC, ws.max_column + 1):
        name = ws.cell(HDR, c).value
        if name is None or str(name).strip() == "":
            continue
        names.append(str(name).strip())
        cols.append(c)
        years.append(ws.cell(YR, c).value)
    return names, years, cols


def _read_slot_rows(ws, HDR, CDAY, CWD, CKIND):
    """枠行を読む。当直列(CKIND)が空でない行を枠とみなし、日/曜日が結合で空なら直前値で補完。"""
    out = []
    prev_day = prev_wd = None
    for r in range(HDR + 1, ws.max_row + 1):
        kind = ws.cell(r, CKIND).value
        if kind is None or str(kind).strip() == "":
            continue
        day = ws.cell(r, CDAY).value
        wd = ws.cell(r, CWD).value
        if day is not None and str(day).strip() != "":
            try:
                day = int(day); prev_day = day
            except (TypeError, ValueError):
                continue  # 見出しの残り等はスキップ
        else:
            day = prev_day
        if wd is not None and str(wd).strip() != "":
            prev_wd = str(wd).strip()
        wd = prev_wd
        if day is None:
            continue
        out.append((r, day, wd or "", str(kind).strip()))
    return out


def solve_roster(file_like, year, month,
                 duty_pref=None,
                 coeff_override=None,
                 holiday_wd_oc=3, holiday_wd_duty=4,
                 oc_single_sat=3, oc_single_sun=2, oc_single_sun_monhol=3, oc_single_holiday=3,
                 top_coeff=2.0,
                 w_premium=80, w_pref=60, w_consec=100, w_dev=10,
                 w_duty_slope=15,
                 w_spacing=120, spacing=None, spacing_cross=None,
                 weekday_prefer=None, w_weekday=40,
                 no_sat_duty_year=None,
                 jr_frisat_year=None, w_frisat=40,
                 senior_duty_cap_year=None, senior_duty_cap=1,
                 time_limit=40):
    wb = load_workbook(file_like)
    ws = wb.active
    HDR, YR, PR, CDAY, CWD, CKIND, FC = _find_layout(ws)
    last_col = ws.max_column

    members, years_raw, mcols = _read_members(ws, HDR, YR, FC)
    NM = len(members)

    # 入局年度の欠損対応：数値化し、欠損は既知年度の中央値で仮設定（index基準）
    def _to_year(v):
        try:
            return int(str(v).strip())
        except (TypeError, ValueError):
            return None
    yrs = [_to_year(v) for v in years_raw]
    valid = [v for v in yrs if v is not None]
    default_year = int(statistics.median(valid)) if valid else 2020
    missing_year = [members[i] for i in range(NM) if yrs[i] is None]
    yrs = [v if v is not None else default_year for v in yrs]

    # 日直/当直希望（index基準）。手動指定(duty_pref)は氏名一致で全該当に適用。
    pref = [None] * NM
    for i in range(NM):
        if PR is not None:
            v = str(ws.cell(PR, mcols[i]).value or "")
            if "日直" in v:
                pref[i] = "day"
            elif "宿直" in v or "当直" in v:
                pref[i] = "night"
        if members[i] in (duty_pref or {}):
            pref[i] = (duty_pref or {})[members[i]]

    slots = []
    for (r, day, wd, k) in _read_slot_rows(ws, HDR, CDAY, CWD, CKIND):
        avail = [i for i in range(NM)
                 if str(ws.cell(r, mcols[i]).value).strip() not in ("×", "X", "x", "✕")]
        slots.append(dict(row=r, day=day, wd=wd, kind=k, avail=avail,
                          nth=(day - 1) // 7 + 1,   # その月で何回目のその曜日か（第N）
                          is_oc=k.startswith("OC"),
                          is_duty=k.startswith("宿直") or k.startswith("日直"),
                          is_nikkyoku=k.startswith("日直")))
    if not slots:
        raise ValueError("枠（データ行）が読み取れませんでした。")

    by_day = {}
    for i, s in enumerate(slots):
        by_day.setdefault(s["day"], []).append(i)
    for day, idxs in by_day.items():
        ocs = [i for i in idxs if slots[i]["is_oc"]]
        if not ocs:
            continue
        wd = slots[ocs[0]]["wd"]
        dt = datetime.date(year, month, day)
        wkend_hol = wd in ("土", "日") or jpholiday.is_holiday(dt)
        for j, i in enumerate(ocs):
            if len(ocs) >= 2:
                slots[i]["oc_role"] = "day" if j == 0 else "night"   # 日中／夜の2枠
            else:
                slots[i]["oc_role"] = "single" if wkend_hol else "weekday"  # 終日1枠／平日
        for i in ocs:
            slots[i]["is_day_oc"] = (slots[i].get("oc_role") == "day")

    def score(s):
        dt = datetime.date(year, month, s["day"]); wd = s["wd"]
        mon_hol = (wd == "日") and jpholiday.is_holiday(dt + datetime.timedelta(days=1))
        if s["is_oc"]:
            role = s.get("oc_role", "weekday")
            if role == "day":                       # 2枠のうち日中OC
                return 3
            if role == "single":                    # 土日祝の終日1枠OC
                if wd == "土": return oc_single_sat
                if wd == "日": return oc_single_sun_monhol if mon_hol else oc_single_sun
                return oc_single_holiday             # 祝日(平日)の終日1枠
            if role == "night":                     # 2枠のうち夜OC
                if wd == "土": return 3
                if wd == "日": return 3 if mon_hol else 2
                if wd == "金": return 2
                if jpholiday.is_holiday(dt): return holiday_wd_oc
                return 2
            # 平日の単独OC
            if wd == "金": return 2
            return 1
        if s["is_nikkyoku"]: return 4
        if wd == "土": return 5
        if wd == "日": return 5 if mon_hol else 3
        if wd == "金": return 3
        if jpholiday.is_holiday(dt): return holiday_wd_duty
        return 2
    for s in slots:
        s["pt"] = score(s)

    def is_premium(s):
        wd, k = s["wd"], s["kind"]
        if wd == "金" and k.startswith("宿直"): return True
        if wd in ("土", "日") and s["is_duty"]: return True
        if wd == "月" and k.startswith("日直"): return True
        if s.get("is_day_oc"): return True
        return False

    model = cp_model.CpModel()
    x = {(si, mi): model.NewBoolVar(f"x_{si}_{mi}")
         for si, s in enumerate(slots) for mi in s["avail"]}

    unfilled = []
    for si, s in enumerate(slots):
        if not s["avail"]:
            unfilled.append(f"{s['day']}日 {s['kind']}")
            continue
        model.Add(sum(x[(si, mi)] for mi in s["avail"]) == 1)
    if unfilled:
        raise ValueError("全員が×で誰も入れない枠があります: " + " / ".join(unfilled))

    for mi in range(NM):
        oc = [x[(si, mi)] for si, s in enumerate(slots) if mi in s["avail"] and s["is_oc"]]
        du = [x[(si, mi)] for si, s in enumerate(slots) if mi in s["avail"] and s["is_duty"]]
        if oc: model.Add(sum(oc) <= 2)
        if du: model.Add(sum(du) <= 2)
        # 指定年度より古い（入局年度が小さい）メンバーは、日直＋当直の合計を上限で制限
        if senior_duty_cap_year is not None and du and yrs[mi] < senior_duty_cap_year:
            model.Add(sum(du) <= senior_duty_cap)

    for day, idxs in by_day.items():
        for mi in range(NM):
            v = [x[(si, mi)] for si in idxs if mi in slots[si]["avail"]]
            if len(v) > 1: model.Add(sum(v) <= 1)

    # 指定年度より古い（入局年度が小さい）メンバーは土曜の当直(宿直)なし。
    # ただし全員が対象で埋められなくなる枠は、実行可能性のため除外を見送る。
    if no_sat_duty_year is not None:
        for si, s in enumerate(slots):
            if s["wd"] == "土" and s["kind"].startswith("宿直"):
                excluded = [mi for mi in s["avail"] if yrs[mi] < no_sat_duty_year]
                allowed = [mi for mi in s["avail"] if mi not in excluded]
                if allowed:  # 割り当て先が残る場合のみ禁止
                    for mi in excluded:
                        model.Add(x[(si, mi)] == 0)

    worked = {}
    for mi in range(NM):
        for day, idxs in by_day.items():
            v = [x[(si, mi)] for si in idxs if mi in slots[si]["avail"]]
            if v:
                w = model.NewBoolVar(f"wk_{mi}_{day}")
                model.Add(sum(v) >= 1).OnlyEnforceIf(w)
                model.Add(sum(v) == 0).OnlyEnforceIf(w.Not())
                worked[(mi, day)] = w
    maxday = max(s["day"] for s in slots)
    consec = []
    for mi in range(NM):
        for d in range(1, maxday):
            a, b = worked.get((mi, d)), worked.get((mi, d + 1))
            if a is not None and b is not None:
                c = model.NewBoolVar(f"cs_{mi}_{d}")
                model.AddBoolAnd([a, b]).OnlyEnforceIf(c)
                model.AddBoolOr([a.Not(), b.Not()]).OnlyEnforceIf(c.Not())
                consec.append(c)

    obj = []
    ymin, ymax = min(yrs), max(yrs)
    coeff = [(coeff_override or {}).get(members[mi],
             seniority_coeff(yrs[mi], ymin, ymax, top_coeff)) for mi in range(NM)]
    total_pt = sum(s["pt"] for s in slots)
    csum = sum(coeff)

    targets = [0] * NM
    for mi in range(NM):
        terms = [s["pt"] * x[(si, mi)] for si, s in enumerate(slots) if mi in s["avail"]]
        pv = model.NewIntVar(0, total_pt, f"pt_{mi}")
        model.Add(pv == sum(terms))
        tgt = round(total_pt * coeff[mi] / csum); targets[mi] = tgt
        dev = model.NewIntVar(0, total_pt, f"dev_{mi}")
        model.Add(dev >= pv - tgt); model.Add(dev >= tgt - pv)
        obj.append(w_dev * dev)

    # 当直/日直の「回数」も入局年度に応じて傾斜（新しい人ほど回数を多く）
    duty_idx = [si for si, s in enumerate(slots) if s["is_duty"]]
    total_duty = len(duty_idx)
    if total_duty and w_duty_slope:
        for mi in range(NM):
            terms = [x[(si, mi)] for si in duty_idx if mi in slots[si]["avail"]]
            dc = model.NewIntVar(0, total_duty, f"dc_{mi}")
            model.Add(dc == (sum(terms) if terms else 0))
            dtgt = round(total_duty * coeff[mi] / csum)
            ddev = model.NewIntVar(0, total_duty, f"ddev_{mi}")
            model.Add(ddev >= dc - dtgt); model.Add(ddev >= dtgt - dc)
            obj.append(w_duty_slope * ddev)

    # 当直・日直の「回数」も入局年度で傾斜（新しい人ほど多く）→ 古い人が多く入る逆転を防ぐ
    if w_duty_slope:
        duty_total = sum(1 for s in slots if s["is_duty"])
        for mi in range(NM):
            dterms = [x[(si, mi)] for si, s in enumerate(slots) if mi in s["avail"] and s["is_duty"]]
            if not dterms:
                continue
            dv = model.NewIntVar(0, duty_total, f"duc_{mi}")
            model.Add(dv == sum(dterms))
            dtgt = round(duty_total * coeff[mi] / csum)
            dd = model.NewIntVar(0, duty_total, f"dudev_{mi}")
            model.Add(dd >= dv - dtgt); model.Add(dd >= dtgt - dv)
            obj.append(w_duty_slope * dd)

    for si, s in enumerate(slots):
        if s["is_duty"]:
            for mi in s["avail"]:
                p = pref[mi]
                if p == "day" and not s["is_nikkyoku"]:
                    obj.append(w_pref * x[(si, mi)])
                if p == "night" and s["is_nikkyoku"]:
                    obj.append(w_pref * x[(si, mi)])

    # 指定メンバー×曜日（任意で週番号・枠種別）を優先（同名は全員に適用）
    for rule in (weekday_prefer or []):
        nm = rule.get("member"); wd = rule.get("weekday")
        kind = rule.get("kind"); weeks = rule.get("weeks")
        for mi in [i for i in range(NM) if members[i] == nm]:
            for si, s in enumerate(slots):
                if mi not in s["avail"] or s["wd"] != wd:
                    continue
                if weeks and s["nth"] not in weeks:
                    continue
                if kind and not s["kind"].startswith(kind):
                    continue
                obj.append(-w_weekday * x[(si, mi)])

    # 指定年度より新しい（入局年度が大きい）メンバーは、金曜・土曜の当直(宿直)を優先（報酬）
    if jr_frisat_year is not None:
        for si, s in enumerate(slots):
            if s["wd"] in ("金", "土") and s["kind"].startswith("宿直"):
                for mi in s["avail"]:
                    if yrs[mi] > jr_frisat_year:
                        obj.append(-w_frisat * x[(si, mi)])

    prem = [si for si, s in enumerate(slots) if is_premium(s)]
    for mi in range(NM):
        v = [x[(si, mi)] for si in prem if mi in slots[si]["avail"]]
        if v:
            cnt = model.NewIntVar(0, len(v), f"pc_{mi}"); model.Add(cnt == sum(v))
            ex = model.NewIntVar(0, len(v), f"px_{mi}"); model.Add(ex >= cnt - 1)
            obj.append(w_premium * ex)

    # 間隔：同じ人の割当が指定日数未満で近接したらペナルティ（同種別＋種別またぎ、1以下=無効）
    def cat_of(s):
        if s["is_oc"]: return "OC"
        if s["kind"].startswith("宿直"): return "宿直"
        if s["kind"].startswith("日直"): return "日直"
        return None
    sp = spacing if spacing is not None else {"宿直": 7, "日直": 7}
    spc = spacing_cross or {}
    def gap_of(ca, cb):
        if ca == cb:
            return sp.get(ca, 0) or 0
        return spc.get(f"{ca}-{cb}", spc.get(f"{cb}-{ca}", 0)) or 0
    for mi in range(NM):
        av = [si for si, s in enumerate(slots) if mi in s["avail"]]
        for a in range(len(av)):
            for b in range(a + 1, len(av)):
                si, sj = av[a], av[b]
                days = gap_of(cat_of(slots[si]), cat_of(slots[sj]))
                if days and days > 1 and abs(slots[si]["day"] - slots[sj]["day"]) < days:
                    near = model.NewBoolVar(f"near_{mi}_{si}_{sj}")
                    model.Add(x[(si, mi)] + x[(sj, mi)] - 1 <= near)
                    obj.append(w_spacing * near)

    for c in consec:
        obj.append(w_consec * c)

    model.Minimize(sum(obj))
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit
    st = solver.Solve(model)
    if st not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError("条件を満たす表が見つかりませんでした。×が多すぎるか、上限が厳しすぎる可能性があります。")

    assign = {si: mi for si, s in enumerate(slots) for mi in s["avail"]
              if solver.Value(x[(si, mi)]) == 1}

    summary = []
    for mi, m in enumerate(members):
        oc = du = pt = pr = 0
        for si, s in enumerate(slots):
            if assign.get(si) == mi:
                oc += s["is_oc"]; du += s["is_duty"]; pt += s["pt"]; pr += is_premium(s)
        summary.append(dict(member=m, year=yrs[mi], coeff=round(coeff[mi], 2),
                            pref={"day": "日直", "night": "当直"}.get(pref[mi], ""),
                            target=targets[mi], points=pt, oc=oc, duty=du, premium=pr))

    for si, mi in assign.items():
        ws.cell(slots[si]["row"], mcols[mi]).value = "◎" if slots[si]["is_duty"] else "〇"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    info = dict(total_points=total_pt, n_slots=len(slots), status=solver.StatusName(st),
                consecutive=sum(solver.Value(c) for c in consec),
                members=members, missing_year=missing_year,
                holidays=[f"{d}日" for d in range(1, maxday + 1)
                          if jpholiday.is_holiday(datetime.date(year, month, d))])
    return buf.getvalue(), summary, info


def template_info(file_like):
    """テンプレExcelから members, years, slots[(row,day,wd,kind)] を返す。"""
    wb = load_workbook(file_like); ws = wb.active
    HDR, YR, PR, CDAY, CWD, CKIND, FC = _find_layout(ws)
    members, years, cols = _read_members(ws, HDR, YR, FC)
    slots = _read_slot_rows(ws, HDR, CDAY, CWD, CKIND)
    return members, years, slots


def apply_photo_grid(file_like, grid):
    """写真から読んだ grid(40x12, 1=×) をテンプレの各メンバー列に書き込み、xlsxのbytesを返す。"""
    wb = load_workbook(file_like); ws = wb.active
    HDR, YR, PR, CDAY, CWD, CKIND, FC = _find_layout(ws)
    members, _years, mcols = _read_members(ws, HDR, YR, FC)
    srows = [r for (r, _d, _w, _k) in _read_slot_rows(ws, HDR, CDAY, CWD, CKIND)]
    for i, r in enumerate(srows):
        if i >= len(grid):
            break
        for j in range(len(mcols)):
            if j >= len(grid[i]):
                break
            ws.cell(r, mcols[j]).value = "×" if grid[i][j] else None
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
