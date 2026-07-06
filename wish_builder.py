"""各科の割り振りExcelから、泌尿器科の当直・オンコール希望表を生成する。
毎月：割り振り(ウロ/整形/皮膚…) → 当直列(宿直N/日直X/OC) に変換して印刷用シートを作る。
"""
import io
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils.datetime import from_excel
import jpholiday

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
FILL_SAT = PatternFill("solid", fgColor="9CC3E6")   # 土=青
FILL_SUN = PatternFill("solid", fgColor="E68AA8")   # 日祝=桃
FILL_DUTY = PatternFill("solid", fgColor="FFF2A8")  # ウロ当直行=黄
MARK = "■"


def _to_date(a):
    if isinstance(a, datetime.datetime):
        return a.date()
    if isinstance(a, datetime.date):
        return a
    return from_excel(a).date()


def read_allocation(alloc_file):
    """割り振りを (date, 曜日, 種別) のリストで返す。"""
    ws = load_workbook(alloc_file, data_only=True).active
    rows = []
    for r in range(1, ws.max_row + 1):
        a, b, c = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
        if isinstance(a, bool):
            continue
        if isinstance(a, (int, float, datetime.date)) and b and c:
            rows.append((_to_date(a), str(b).strip(), str(c).strip()))
    rows.sort(key=lambda t: t[0])
    return rows


def build_slots(alloc_rows, uro_prefix="ウロ"):
    """割り振り → 当直行 [(day or None, 曜日 or None, kind)] に変換。"""
    slots = []
    num = 0        # 平日ウロ当直の連番（宿直N）
    letter = 0     # 土日祝ウロ当直の連番（日直X/宿直X）
    for (d, wd, typ) in alloc_rows:
        holiday = d.weekday() >= 5 or jpholiday.is_holiday(d)
        is_uro = typ.startswith(uro_prefix)
        if is_uro and holiday:
            L = chr(ord("A") + letter); letter += 1
            slots.append((d.day, wd, f"日直{L}"))
            slots.append((None, None, f"宿直{L}"))
        elif is_uro:
            num += 1
            slots.append((d.day, wd, f"宿直{num}"))
        elif holiday:
            slots.append((d.day, wd, "OC"))
            slots.append((None, None, "OC"))
        else:
            slots.append((d.day, wd, "OC"))
    return slots


def read_roster(template_file):
    """ベーステンプレから氏名・入局年度・（あれば）日直希望を読む。"""
    from duty_solver import _find_layout, _read_members
    ws = load_workbook(template_file).active
    HDR, YR, PR, CDAY, CWD, CKIND, FC = _find_layout(ws)
    members, years, cols = _read_members(ws, HDR, YR, FC)
    prefs = []
    if PR is not None:
        for c in cols:
            prefs.append(ws.cell(PR, c).value)
    else:
        prefs = [None] * len(members)
    return members, years, prefs


def build_wish_sheet(alloc_file, template_file, uro_prefix="ウロ"):
    """割り振り + ベーステンプレ → 希望表(.xlsx) のbytesを生成。"""
    alloc = read_allocation(alloc_file)
    if not alloc:
        raise ValueError("割り振りから日付データを読めませんでした（A列=日付, B列=曜日, C列=担当科）。")
    slots = build_slots(alloc, uro_prefix)
    members, years, prefs = read_roster(template_file)
    NM = len(members)
    year, month = alloc[0][0].year, alloc[0][0].month
    holidays = sorted({d.day for (d, _w, _t) in alloc
                       if jpholiday.is_holiday(d)})

    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    FC = 4                      # 氏名開始列（D）
    last_col = FC + NM - 1      # 最終氏名列
    mark_col = last_col + 1     # 右側マーク列

    # ---- ヘッダー ----
    ws.cell(1, 1, MARK); ws.cell(1, mark_col, MARK)
    ws.cell(1, 4, f"　{month}月当直・オンコール　アンケート").font = Font(bold=True, size=12)
    ws.cell(2, 1, "ＬＡＳＴの方は野口先生までお願いします。")
    ws.cell(2, 10, "(締め切り　  月    日  AM）").font = Font(color="C00000")
    ws.cell(4, 1, "医　師　名")
    ws.cell(5, 3, "入局年度")
    for i, y in enumerate(years):
        ws.cell(5, FC + i, y).alignment = CENTER
    ws.cell(6, 1, "日"); ws.cell(6, 2, "曜日"); ws.cell(6, 3, "当直")
    for i, m in enumerate(members):
        ws.cell(6, FC + i, m)
    for c in range(1, last_col + 1):
        cell = ws.cell(6, c); cell.font = Font(bold=True); cell.alignment = CENTER; cell.border = BORDER

    # ---- データ ----
    r0 = 7
    for idx, (day, wd, kind) in enumerate(slots):
        r = r0 + idx
        if day is not None:
            ws.cell(r, 1, day).alignment = CENTER
            ws.cell(r, 2, wd).alignment = CENTER
        ws.cell(r, 3, kind).alignment = CENTER
        for c in range(1, last_col + 1):
            ws.cell(r, c).border = BORDER
        # 土日祝の色（先頭行のみ日付があるので、その色を両行に付ける）
        is_dutyrow = kind.startswith("宿直") or kind.startswith("日直")
        # 曜日は結合下段だと空なので、先頭行の曜日から判定
        wd_here = wd if wd else (slots[idx - 1][1] if idx > 0 else None)
        day_here = day if day is not None else (slots[idx - 1][0] if idx > 0 else None)
        hol = False
        if day_here is not None:
            try:
                hol = jpholiday.is_holiday(datetime.date(year, month, day_here))
            except ValueError:
                hol = False
        if wd_here == "土":
            ws.cell(r, 1).fill = ws.cell(r, 2).fill = FILL_SAT
        elif wd_here == "日" or hol:
            ws.cell(r, 1).fill = ws.cell(r, 2).fill = FILL_SUN
        if is_dutyrow:
            for c in range(3, last_col + 1):   # 当直〜メンバー記入枠まで黄色
                ws.cell(r, c).fill = FILL_DUTY

    last_row = r0 + len(slots) - 1

    # 土日祝の2行（日/曜日）を結合
    i = 0
    while i < len(slots):
        if i + 1 < len(slots) and slots[i + 1][0] is None:  # 次行が空=結合ペア
            rr = r0 + i
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr + 1, end_column=1)
            ws.merge_cells(start_row=rr, start_column=2, end_row=rr + 1, end_column=2)
            i += 2
        else:
            i += 1

    # 下部の四隅マークと注意書き
    ws.cell(last_row + 1, 1, MARK); ws.cell(last_row + 1, mark_col, MARK)
    ws.cell(last_row + 3, 1,
            "×のみ記入してください。○を記入しないでください。"
            "学会などの休みで矢印を入れないでください。読み取れなくなります。")

    # 列幅
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 8
    for i in range(NM):
        ws.column_dimensions[ws.cell(6, FC + i).column_letter].width = 6

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    info = dict(year=year, month=month, n_slots=len(slots), members=members,
                holidays=[f"{d}日" for d in holidays])
    return buf.getvalue(), info
